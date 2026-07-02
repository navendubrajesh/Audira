#!/usr/bin/env python3
"""Sync backlog-data.json from Enterprise_Comms_NeuroAnalyzer_Competitive_Backlog.xlsx."""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parents[1]
XLSX = REPO / "Enterprise_Comms_NeuroAnalyzer_Competitive_Backlog.xlsx"
OUT = ROOT / "src/mock/backlog-data.json"


def rows_to_dicts(ws):
    headers = [c for c in next(ws.iter_rows(values_only=True))]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).startswith("TCA"):
            continue
        story_id = str(row[0])
        num = int(story_id.split("-")[1])
        if num > 92:
            continue
        items.append({h: ("" if v is None else v) for h, v in zip(headers, row)})
    return items


def main():
    if not XLSX.exists():
        print(f"Missing workbook: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    backlog = rows_to_dicts(wb["Backlog"])

    ws = wb["Competitor Landscape"]
    headers = [c for c in next(ws.iter_rows(values_only=True))]
    competitors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            competitors.append({h: ("" if v is None else v) for h, v in zip(headers, row)})

    payload = {"backlog": backlog, "competitors": competitors}
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Synced {len(backlog)} stories + {len(competitors)} competitors -> {OUT}")


if __name__ == "__main__":
    main()
